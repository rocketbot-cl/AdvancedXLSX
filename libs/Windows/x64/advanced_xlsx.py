from bs4 import BeautifulSoup, element
from openpyxl import Workbook, worksheet
import xlrd
from openpyxl.utils.cell import column_index_from_string
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from datetime import datetime, timedelta
import csv

class AdvancedXlsx:

    def __init__(self, wb=Workbook(), sheet=None):
        self.wb = wb
        if sheet is None:
           sheet = self.wb.active
        self.sheet = sheet

    def open_xls(self, path: str, col = None)->Workbook:
        
        from tablepyxl import tablepyxl
        
        # if platform.system() == 'Windows':
        #     fname = path
        #     excel = win32.gencache.EnsureDispatch('Excel.Application')
        #     excel.Visible = False
        #     wb = excel.Workbooks.Open(fname)

        #     wb.SaveAs(fname+"x", FileFormat = 51)   #FileFormat = 51 is for .xlsx extension
        #     wb.Close()                              #FileFormat = 56 is for .xls extension
        #     excel.Application.Quit()
            
            # https://stackoverflow.com/questions/9918646/how-to-convert-xls-to-xlsx

        try:
            # open html, checks that there is a table in it and convert to xlsx with tablepyxl 
            with open(path, "r", encoding="latin-1") as f:
                table = f.read()
                soup = BeautifulSoup(table, 'html.parser')
                if soup.find_all("table"):
                    self.wb = tablepyxl.document_to_workbook(table)
                    return self.wb
                
            # open csv and convert to xlsx with openpyxl    
            with open(path, "r", encoding="latin-1") as f:
                reader = csv.reader(f, delimiter="\t")
                if reader:            
                    self.wb = Workbook()
                    ws = self.wb.active
                    for row in reader:
                        ws.append(row)
                    return self.wb
        except:
            self.convert_xls(path, Workbook(), col)
            return self.wb

        
    def convert_xls(self, path:str, wb, col = None):
        self.wb = wb
        self.sheet = self.wb.active
        wb_ = xlrd.open_workbook(path)
        sheets = wb_.sheet_names()
        
        if len(sheets) == 1:
            sheet_ = wb_.sheet_by_name(sheets[0])
            
            for i in range(sheet_.nrows):
                row = [sheet_.cell_value(rowx=i, colx=j) for j in range(sheet_.ncols)]
                
                row = list(map(lambda x: str(x).encode('latin-1',errors='ignore').decode() if not isinstance(x,int) and not isinstance(x,float) else x, row))
                
                # Format data as date for the columns given
                if col:
                    
                    for c in col:
                        c = eval(c)
                        try:
                            row[c] = xlrd.xldate_as_datetime(row[c], 0).date().strftime("%d-%m-%Y")
                        except:
                            print(f"Data in row {i} - col {c} is not a number.")
                            
                self.sheet.append(row)
                self.sheet.title = sheets[0]
                
        if len(sheets) > 1:
            for sheet in sheets:
                self.wb.create_sheet(sheet)
                sheet_ = wb_.sheet_by_name(sheet)
                for i in range(sheet_.nrows):
                    row = [sheet_.cell_value(rowx=i, colx=j) for j in range(sheet_.ncols)]
                    
                    # Format data as date for the columns given
                    if col:
                        for c in col:
                            c = eval(c)
                        try:
                            row[c] = xlrd.xldate_as_datetime(row[c], 0).date().strftime("%d-%m-%Y")
                        except:
                            print(f"Data in row {i} - col {c} is not a number.")
                    
                    self.wb[sheet].append(row)
           
            # It deletes the default sheet, because the loop already creates one for one in the xls
            del self.wb[self.sheet.title]
    
    def convert_to_csv(self, path_csv: str, date_format: str, delimiter: str = ",")->None:
        with open(path_csv, "w", newline="", encoding="utf-8") as f:
            c = csv.writer(f, delimiter=delimiter)
            for r in self.sheet.rows:
                row = []
                for cell in r:
                    if isinstance(cell.value, datetime):
                        formatted_date = cell.value.strftime(date_format)
                        row.append(formatted_date)
                    elif isinstance(cell.value, float) and cell.value.is_integer():
                        row.append(int(cell.value))
                    else:
                        row.append(cell.value)
                c.writerow(row)
    
    
    
    # Deprecated after using tablepyxl library
    def get_from_html(self, soup: BeautifulSoup)->None:
        tables = soup.findChildren()
        for table in tables:
            self.get_table(table)
    # Deprecated after using tablepyxl library            
    def get_table(self, table: element.Tag)->None:
        count = 1
        for row in table.findChildren("tr", recursive=False):
            
            for j, col in enumerate(row.findChildren("td", recursive=False), start=1):
                if col.table:
                    i = count
                    for ii, sub_row in enumerate(col.table.findChildren("tr", recursive=False)):
                        for jj, sub_col in enumerate(sub_row.findChildren("td", recursive=False), start=1):
                            self.sheet.cell(row=i+ii, column=j).value = sub_col.text.strip()
                else:
                    self.sheet.cell(row=count, column=j).value = col.text.strip()
            
            for k, col in enumerate(row.findChildren("th", recursive=False), start=1):
                if col.table:
                    o = count
                    for oo, sub_row in enumerate(col.table.findChildren("tr", recursive=False)):
                        for kk, sub_col in enumerate(sub_row.findChildren("th", recursive=False), start=1):
                            self.sheet.cell(row=o+oo, column=k).value = sub_col.text.strip()
                else:
                    self.sheet.cell(row=count, column=k).value = col.text.strip()
            count = self.sheet.max_row + 1

    def change_sheet(self, sheetname: str)->worksheet:
        self.sheet = self.wb.get_sheet_by_name(sheetname)
        return self.sheet

    def insert_columns(self, col_range: str)->None:
        if ":" in col_range:
            col_range = col_range.split(":")
            start = column_index_from_string(col_range[0])
            try:
                end = column_index_from_string(col_range[1]) + 1
            except IndexError:
                end = start + 1
            self.sheet.insert_cols(start, end - start)
        else:
            column = column_index_from_string(col_range)
            self.sheet.insert_cols(column)
    
    def insert_rows(self, row_range: str)->None:
        
        if ":" in row_range:
            a = int(row_range.split(":")[0])
            b = int(row_range.split(":")[1])
            self.sheet.insert_rows(int(a), int(b))
        else:
            self.sheet.insert_rows(int(row_range))
    
    def delete_columns(self, col_range: str)->None:
        
        if ":" in col_range:
            col_range = col_range.split(":")
            start = column_index_from_string(col_range[0])
            try:
                end = column_index_from_string(col_range[1]) + 1
            except IndexError:
                end = start + 1
            self.sheet.delete_cols(start, end - start)
        else:
            column = column_index_from_string(col_range)
            self.sheet.delete_cols(column)
            
    def delete_rows(self, row_range: str)->None:
        
        if ":" in row_range:
            a = int(row_range.split(":")[0])
            b = int(row_range.split(":")[1])
            self.sheet.delete_rows(int(a), int(b))
        else:
            self.sheet.delete_rows(int(row_range))

    def count_by_range(self, range_:str)->tuple:
        column = self.sheet[range_].column
        row = self.sheet[range_].row
        col_length = len([column for column in self.sheet.columns][column - 1])
        row_length = len([row for row in self.sheet.rows][row - 1])

        return (col_length, row_length)
    
    def get_cells_by_filter(self, filter_:list, only_data=True)->list:
        raise NotImplementedError

    def change_format(self, sheet, range, format=None, h=None, v=None):
        datas = self.wb[sheet][range]
        
        for data in datas:
            if isinstance(data, tuple):
                for d in data:
                    if d:
                        # DO not erase the next line. It is necessary to make available '_style.numFmtId' to change the format.
                        if format:
                            d.style_id
                            d._style.numFmtId = int(format)
                        if h or v:
                            d.alignment = Alignment(horizontal=h, vertical=v)
            else:
                if format:
                    d.style_id
                    d._style.numFmtId = int(format)
                if h or v:
                    d.alignment = Alignment(horizontal=h, vertical=v)
    
    global get_columns_between
    def get_columns_between(start_col, end_col):
        start_num = ord(start_col.upper()) - 64
        end_num = ord(end_col.upper()) - 64
        columns = []
        for i in range(start_num, end_num + 1):
            columns.append(chr(i + 64))
        return columns
    
    def change_format_col(self, sheet, cols, format=None):
        # Can't be used for alignment formatting
        sheet_ = self.wb[sheet]
        columns = get_columns_between(cols[0], cols[1])
    
        for column in columns:
            column_ = sheet_.column_dimensions[column]
            if format:
                column_.style_id
                column_._style.numFmtId = int(format)
    
    def change_format_row(self, sheet, row, format=None):
        # Can't be used for alignment formatting
        sheet_ = self.wb[sheet]
        rows = [i for i in range(row[0], row[1]+1)]
        for row in rows:
            row_ = sheet_.row_dimensions[row]
            if format:
                row_.style_id
                row_._style.numFmtId = int(format) 

    
    def new_sheet(self, sheet):
        self.wb.create_sheet(sheet)
    
    def del_sheet(self, sheet):
        self.wb.remove(sheet)
    
    def insert_image(self, sheet, path, cell):
        img = Image(path)
        self.wb[sheet].add_image(img, cell)
        
    def read_range(self, sheet, range):
        return self.wb[sheet][range]
    
    @staticmethod
    def get_excel_date(date_time_str):
        import datetime
        
        UTC = datetime.timezone.utc
        dt_obj = datetime.datetime.fromisoformat(date_time_str).replace(tzinfo=UTC)
        day_zero = datetime.datetime(1899,12,30, tzinfo=UTC)

        excel_serial_date = (dt_obj-day_zero).total_seconds()/86400

        return excel_serial_date