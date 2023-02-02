from bs4 import BeautifulSoup, element
from openpyxl import Workbook, worksheet, load_workbook
import xlrd
from openpyxl.utils.cell import column_index_from_string
import pandas as pd
import win32com.client as win32
import platform
import csv

class AdvancedXlsx:

    def __init__(self, wb=Workbook(), sheet=None):
        self.wb = wb
        if sheet is None:
           sheet = self.wb.active
        self.sheet = sheet

    def open_xls(self, path: str, col = None)->Workbook:
        
        from tablepyxl import tablepyxl
        
        # if platform.system() == 'Windows':
        #     fname = path
        #     excel = win32.gencache.EnsureDispatch('Excel.Application')
        #     excel.Visible = False
        #     wb = excel.Workbooks.Open(fname)

        #     wb.SaveAs(fname+"x", FileFormat = 51)   #FileFormat = 51 is for .xlsx extension
        #     wb.Close()                              #FileFormat = 56 is for .xls extension
        #     excel.Application.Quit()
            
            # https://stackoverflow.com/questions/9918646/how-to-convert-xls-to-xlsx

        try:
            # open html, checks that there is a table in it and convert to xlsx with tablepyxl 
            with open(path, "r", encoding="latin-1") as f:
                table = f.read()
                soup = BeautifulSoup(table, 'html.parser')
                if soup.find_all("table"):
                    self.wb = tablepyxl.document_to_workbook(table)
                    return self.wb
                
            # open csv and convert to xlsx with openpyxl    
            with open(path, "r", encoding="latin-1") as f:
                reader = csv.reader(f, delimiter="\t")
                print("csv: ", reader)
                if reader:            
                    self.wb = Workbook()
                    ws = self.wb.active
                    for row in reader:
                        ws.append(row)
                    return self.wb
        except:
            self.convert_xls(path, col)
            return self.wb

        
    def convert_xls(self, path:str, col = None)->None:
        self.wb = Workbook()
        self.sheet = self.wb.active
        wb = xlrd.open_workbook(path)
        sheets = wb.sheet_names()
        
        if len(sheets) == 1:
            sheet_ = wb.sheet_by_name(sheets[0])
            
            for i in range(sheet_.nrows):
                row = [sheet_.cell_value(rowx=i, colx=j) for j in range(sheet_.ncols)]
                
                # Format data as date for the columns given
                if col:
                    for c in col:
                        c = eval(c)
                        try:
                            row[c] = xlrd.xldate_as_datetime(row[c], 0).date().strftime("%d-%m-%Y")
                        except:
                            print(f"Data in row {i} - col {c} is not a number.")
                self.sheet.append(row)
                self.sheet.title = sheets[0]
                
        if len(sheets) > 1:
            for sheet in sheets:
                self.wb.create_sheet(sheet)
                sheet_ = wb.sheet_by_name(sheet)
                for i in range(sheet_.nrows):
                    row = [sheet_.cell_value(rowx=i, colx=j) for j in range(sheet_.ncols)]
                    
                    # Format data as date for the columns given
                    if col:
                        for c in col:
                            c = eval(c)
                        try:
                            row[c] = xlrd.xldate_as_datetime(row[c], 0).date().strftime("%d-%m-%Y")
                        except:
                            print(f"Data in row {i} - col {c} is not a number.")
                    
                    self.wb[sheet].append(row)
           
            # It deletes the default sheet, because the loop already creates one for one in the xls
            del self.wb[self.sheet.title]
        
    # Deprecated after using tablepyxl library
    def get_from_html(self, soup: BeautifulSoup)->None:
        tables = soup.findChildren()
        for table in tables:
            self.get_table(table)
    # Deprecated after using tablepyxl library            
    def get_table(self, table: element.Tag)->None:
        count = 1
        for row in table.findChildren("tr", recursive=False):
            
            for j, col in enumerate(row.findChildren("td", recursive=False), start=1):
                if col.table:
                    i = count
                    for ii, sub_row in enumerate(col.table.findChildren("tr", recursive=False)):
                        for jj, sub_col in enumerate(sub_row.findChildren("td", recursive=False), start=1):
                            self.sheet.cell(row=i+ii, column=j).value = sub_col.text.strip()
                else:
                    self.sheet.cell(row=count, column=j).value = col.text.strip()
            
            for k, col in enumerate(row.findChildren("th", recursive=False), start=1):
                if col.table:
                    o = count
                    for oo, sub_row in enumerate(col.table.findChildren("tr", recursive=False)):
                        for kk, sub_col in enumerate(sub_row.findChildren("th", recursive=False), start=1):
                            self.sheet.cell(row=o+oo, column=k).value = sub_col.text.strip()
                else:
                    self.sheet.cell(row=count, column=k).value = col.text.strip()
            count = self.sheet.max_row + 1

    def change_sheet(self, sheetname: str)->worksheet:
        self.sheet = self.wb.get_sheet_by_name(sheetname)
        return self.sheet

    def insert_columns(self, col_range: str)->None:
        if ":" in col_range:
            col_range = col_range.split(":")
            start = column_index_from_string(col_range[0])
            try:
                end = column_index_from_string(col_range[1]) + 1
            except IndexError:
                end = start + 1
            self.sheet.insert_cols(start, end - start)
        else:
            column = column_index_from_string(col_range)
            self.sheet.insert_cols(column)
    
    def insert_rows(self, row_range: str)->None:
        
        if ":" in row_range:
            a = int(row_range.split(":")[0])
            b = int(row_range.split(":")[1])
            self.sheet.insert_rows(int(a), int(b))
        else:
            self.sheet.insert_rows(int(row_range))
    
    def delete_columns(self, col_range: str)->None:
        
        if ":" in col_range:
            col_range = col_range.split(":")
            start = column_index_from_string(col_range[0])
            try:
                end = column_index_from_string(col_range[1]) + 1
            except IndexError:
                end = start + 1
            self.sheet.delete_cols(start, end - start)
        else:
            column = column_index_from_string(col_range)
            self.sheet.delete_cols(column)
            
    def delete_rows(self, row_range: str)->None:
        
        if ":" in row_range:
            a = int(row_range.split(":")[0])
            b = int(row_range.split(":")[1])
            self.sheet.delete_rows(int(a), int(b))
        else:
            self.sheet.delete_rows(int(row_range))

    def count_by_range(self, range_:str)->tuple:
        column = self.sheet[range_].column
        row = self.sheet[range_].row
        col_length = len([column for column in self.sheet.columns][column - 1])
        row_length = len([row for row in self.sheet.rows][row - 1])

        return (col_length, row_length)
    
    def get_cells_by_filter(self, filter_:list, only_data=True)->list:
        raise NotImplementedError

    def change_format(self, sheet, range, format):
        datas = self.wb[sheet][range]
        
        for data in datas:
            if isinstance(data, tuple):
                for d in data:
                    if d:
                        # DO not erase the next line. It is necessary to make available '_style.numFmtId' to change the format.
                        d.style_id
                        d._style.numFmtId = int(format)
            else:
                d.style_id
                d._style.numFmtId = int(format)
    
    def new_sheet(self, sheet):
        self.wb.create_sheet(sheet)
    
    def del_sheet(self, sheet):
        self.wb.remove(sheet)